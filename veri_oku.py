# -*- coding: utf-8 -*-
"""
veri_oku.py  --  Beko Cerkezkoy TV Anakart Cizelgeleme / Veri Okuma Katmani
---------------------------------------------------------------------------
Bu modul, "Sasi_Uretim_Plani_v4_1.xlsx" dosyasini okur ve MILP modelinin
ihtiyac duydugu SABIT GIRDILERI Python sozluklerine cevirir.

Modelin karar degiskenleri (gunluk uretim, kumulatif uretim, kalan stok)
BURADA OKUNMAZ -- onlari optimizasyon modeli kendisi uretir.

Planlama ufku : 14 is gunu  (7 Mayis 2026 - 23 Mayis 2026)  -> Excel sutun D..Q
Onemli not    : Excel'de R sutunu 23 Mayis'in SEHVEN tekrari oldugu icin
                veri okuma sirasinda atilir; ufuk 14 gune sabitlenir.
"""

from pathlib import Path
import openpyxl


# ===========================================================================
# 1. SABITLER  --  Excel dosyasinin gercek yerlesimine gore tanimlanmistir
# ===========================================================================

# Planlama ufku: Excel sutun indeksleri (1=A, 2=B, ...). D=4 ... Q=17 -> 14 gun.
# R=18 BILEREK haric birakildi (23 Mayis tekrari).
ILK_GUN_SUTUN = 4    # D
SON_GUN_SUTUN = 17   # Q
UFUK_GUN_SAYISI = SON_GUN_SUTUN - ILK_GUN_SUTUN + 1   # = 14

# 14 kart (Stok sayfasindaki resmi sira)
KARTLAR = ['F4', 'GB', 'GL', 'GX', 'LG', 'MR', 'V1',
           'XC', 'XD', 'XGB', 'XGS', 'XR', 'Y3', 'Y4']

# SUS sayfasi blok satir araliklari  (satir no, 1-indeksli)
SUS_BLOKLARI = {
    'otd_alokasyon':  (4, 15),     # OTD Hat-Kart Alokasyonu  (hat satirlari)
    'otd_oran':       (17, 28),    # OTD Oranlar (kapasite dusus katsayisi)
    'md_alokasyon':   (86, 91),    # MD Hat-Kart Alokasyonu
    'md_oran':        (93, 98),    # MD Oranlar
    'ta_fikstur':     (170, 183),  # TA Fikstur Kullanimi (kart bazli)
    'baslangic_stok': None,        # Stok ayri sayfadan okunur
    'montaj_plani':   (250, 263),  # TV Montaj Plani (kart x gun)
    'kadro':          (285, 306),  # SUS icindeki donem kadrosu
}


# ===========================================================================
# 2. YARDIMCI FONKSIYONLAR
# ===========================================================================

def _normalize_kart(ad):
    """Kart isimlerini standart hale getirir (XG-B -> XGB gibi)."""
    if ad is None:
        return None
    ad = str(ad).strip().upper().replace('-', '').replace(' ', '')
    return ad


def _sayi(deger, varsayilan=0.0):
    """Hucre degerini guvenli sekilde sayiya cevirir; bos ise varsayilan."""
    if deger is None or deger == '':
        return varsayilan
    try:
        return float(deger)
    except (TypeError, ValueError):
        return varsayilan


# ===========================================================================
# 3. ANA OKUMA FONKSIYONU
# ===========================================================================

def veri_oku(dosya_yolu):
    """
    Excel dosyasini okur ve modelin sabit girdilerini iceren bir sozluk dondurur.

    Dondurulen sozlugun anahtarlari:
      kartlar          : list[str]                       -- 14 kart
      gunler           : list[int]                       -- [1..14]
      gun_tarih        : dict{gun -> 'YYYY-MM-DD'}        -- gun -> takvim tarihi
      otd_hatlari      : list[str]                       -- SUS'ta fiilen kullanilan OTD hatlari
      md_hatlari       : list[str]                       -- SUS'ta fiilen kullanilan MD hatlari
      md_kartlari      : set[str]                         -- MD asamasina giren kartlar (I_MD)
      atla_kartlari    : set[str]                         -- MD'yi atlayip dogrudan TA'ya gidenler (I_atla)
      otd_alokasyon    : dict{(hat,gun) -> kart}          -- o gun o hatta uretilecek kart
      otd_oran         : dict{(hat,gun) -> float}         -- OTD kapasite dusus katsayisi
      md_alokasyon     : dict{(hat,gun) -> kart}
      md_oran          : dict{(hat,gun) -> float}
      tempo            : dict{(hat,kart) -> float}        -- OTD/MD tempolari (UPM)
      fikstur          : dict{(kart,gun) -> int|None}      -- TA gunluk fikstur sayisi
                                                              (None = o gune plan girilmemis)
      fikstur_baz      : dict{kart -> int}                 -- TA kart bazli baz fikstur (SUS C sutunu)
      adet_vardiya     : dict{kart -> float}              -- TA bir fikstur/vardiya uretim adedi
      test_sistemi     : dict{kart -> str}                -- kartin TA test sistemi
      baslangic_otd    : dict{kart -> float}              -- OTD devreden stok
      baslangic_md     : dict{kart -> float}              -- MD devreden stok
      baslangic_ta     : dict{kart -> float}              -- TA devreden stok
      montaj_plani     : dict{(kart,gun) -> float}        -- gunluk son montaj talebi (D..Q)
      kadro            : dict{hat -> int}                 -- donem kadrosu (0 ise hat kapali)
    """
    dosya_yolu = Path(dosya_yolu)
    if not dosya_yolu.exists():
        raise FileNotFoundError(f"Excel dosyasi bulunamadi: {dosya_yolu}")

    wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
    sus = wb['SUS']

    veri = {}
    veri['kartlar'] = list(KARTLAR)
    veri['gunler'] = list(range(1, UFUK_GUN_SAYISI + 1))

    # --- 3.1  Gun -> tarih eslemesi (SUS satir 2, sutun D..Q) ---------------
    gun_tarih = {}
    for g, c in enumerate(range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1), start=1):
        t = sus.cell(row=2, column=c).value
        gun_tarih[g] = t.strftime('%Y-%m-%d') if hasattr(t, 'strftime') else str(t)
    veri['gun_tarih'] = gun_tarih

    # --- 3.2  OTD alokasyon + oran -----------------------------------------
    otd_alokasyon, otd_oran, otd_hatlari = {}, {}, []
    a0, a1 = SUS_BLOKLARI['otd_alokasyon']
    o0, _ = SUS_BLOKLARI['otd_oran']
    for i, r in enumerate(range(a0, a1 + 1, 2)):     # her hat 2 satir, ust satiri al
        hat = sus.cell(row=r, column=2).value        # B sutunu = hat adi
        if hat is None:
            continue
        hat = str(hat).strip()
        kullanildi = False
        r_oran = o0 + i * 2                          # oran blogunda karsilik gelen satir
        for g, c in enumerate(range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1), start=1):
            kart = _normalize_kart(sus.cell(row=r, column=c).value)
            if kart:
                otd_alokasyon[(hat, g)] = kart
                otd_oran[(hat, g)] = _sayi(sus.cell(row=r_oran, column=c).value, 1.0)
                kullanildi = True
        if kullanildi:
            otd_hatlari.append(hat)
    veri['otd_alokasyon'] = otd_alokasyon
    veri['otd_oran'] = otd_oran
    veri['otd_hatlari'] = otd_hatlari

    # --- 3.3  MD alokasyon + oran  (KANAL TEMELLI -- Yol A) ----------------
    # MD blogunda her fiziksel hat 3 satir kaplar (MD1: 86-88, MD2: 89-91).
    # ONEMLI: Bir fiziksel hat ayni gun birden fazla kart isleyebilir
    # (orn. MD2 = XGB + Y4 paralel). Her DOLU satir ayri bir "kanal"dir.
    # Bir hatta tek kanal varsa kanal adi = hat adi (orn. 'MD1').
    # Birden fazla kanal varsa '<hat>_1', '<hat>_2', ... olarak numaralanir.
    # md_kanal_hat : kanal -> ait oldugu fiziksel hat  (paylasimli kapasite icin)
    md_alokasyon, md_oran, md_hatlari = {}, {}, []
    md_kanal_hat = {}
    m0, m1 = SUS_BLOKLARI['md_alokasyon']
    mo0, _ = SUS_BLOKLARI['md_oran']

    # Once her fiziksel hattin satir araligini topla
    hat_satirlari = {}          # hat -> [satir, satir, ...]
    for r in range(m0, m1 + 1):
        hat = sus.cell(row=r, column=2).value
        if hat is None:
            continue
        hat_satirlari.setdefault(str(hat).strip(), []).append(r)

    # Her hat icin yalniz DOLU satirlari kanal olarak ac
    for hat, satirlar in hat_satirlari.items():
        dolu_satirlar = []
        for r in satirlar:
            dolu = any(_normalize_kart(sus.cell(row=r, column=c).value)
                       for c in range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1))
            if dolu:
                dolu_satirlar.append(r)
        tek_kanal = (len(dolu_satirlar) <= 1)
        for k_idx, r in enumerate(dolu_satirlar, start=1):
            kanal = hat if tek_kanal else f"{hat}_{k_idx}"
            r_oran = mo0 + (r - m0)          # oran blogunda karsilik gelen satir
            for g, c in enumerate(range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1), start=1):
                kart = _normalize_kart(sus.cell(row=r, column=c).value)
                if kart:
                    md_alokasyon[(kanal, g)] = kart
                    md_oran[(kanal, g)] = _sayi(sus.cell(row=r_oran, column=c).value, 1.0)
            md_hatlari.append(kanal)
            md_kanal_hat[kanal] = hat
    veri['md_alokasyon'] = md_alokasyon
    veri['md_oran'] = md_oran
    veri['md_hatlari'] = md_hatlari        # ARTIK KANAL listesi (MD1, MD2_1, MD2_2)
    veri['md_kanal_hat'] = md_kanal_hat    # kanal -> fiziksel hat eslemesi

    # --- 3.4  Tempolar sayfasi  (hat-kart tempo, adet/vardiya) -------------
    # NOT: Fikstur sayisi BURADAN OKUNMAZ. Fikstur, SUS sayfasinda kart x gun
    # matrisi olarak verildigi icin asagida 3.4b blogunda okunur.
    tp = wb['Tempolar']
    tempo, adet_vardiya, test_sistemi = {}, {}, {}
    # Tempolar basliklari satir 2: C..I = OD0..OD7, J..K = MD1,MD2
    hat_sutunlari = {}
    for c in range(3, 12):
        h = tp.cell(row=2, column=c).value
        if h:
            hat_sutunlari[str(h).strip()] = c
    for r in range(3, 17):
        kart = _normalize_kart(tp.cell(row=r, column=2).value)
        if not kart:
            continue
        for hat, c in hat_sutunlari.items():
            v = tp.cell(row=r, column=c).value
            if v is not None and v != '':
                tempo[(hat, kart)] = _sayi(v)
        adet_vardiya[kart] = _sayi(tp.cell(row=r, column=14).value)
        ts = tp.cell(row=r, column=12).value
        test_sistemi[kart] = str(ts).strip() if ts else None
    veri['tempo'] = tempo
    veri['adet_vardiya'] = adet_vardiya
    veri['test_sistemi'] = test_sistemi

    # --- 3.4b  TA fikstur  (SUS satir 170-183) -- KART x GUN MATRISI -------
    # SUS'ta her kartin TA fikstur sayisi gun gun verilir (sutun D..Q).
    #   - Gunluk hucre DOLU ise   -> o gunun fikstur sayisi.
    #   - Gunluk hucre BOS ise    -> o gune TA plani henuz GIRILMEMIS demektir;
    #     deger None birakilir, modelin onarim/planlama asamasi doldurur.
    #   - C sutunu (sutun 3) kart bazli bir BAZ degerdir; matris None oldugunda
    #     basvurulabilecek referans olarak ayrica saklanir.
    tf0, tf1 = SUS_BLOKLARI['ta_fikstur']
    fikstur = {}            # (kart,gun) -> int  ya da  None (plan girilmemis)
    fikstur_baz = {}        # kart -> C sutunundaki baz fikstur
    for r in range(tf0, tf1 + 1):
        kart = _normalize_kart(sus.cell(row=r, column=2).value)
        if not kart or kart not in KARTLAR:
            continue
        fikstur_baz[kart] = int(_sayi(sus.cell(row=r, column=3).value))
        for g, c in enumerate(range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1), start=1):
            v = sus.cell(row=r, column=c).value
            fikstur[(kart, g)] = int(_sayi(v)) if v not in (None, '') else None
    veri['fikstur'] = fikstur
    veri['fikstur_baz'] = fikstur_baz

    # --- 3.5  Process_Mapping  ->  MD'ye giren / atlayan kart kumeleri ------
    pm = wb['Process_Mapping']
    md_kartlari, atla_kartlari = set(), set()
    for r in range(2, 16):
        kart = _normalize_kart(pm.cell(row=r, column=2).value)
        if not kart:
            continue
        md_durum = pm.cell(row=r, column=4).value
        if md_durum and str(md_durum).strip().lower() == 'var':
            md_kartlari.add(kart)
        else:
            atla_kartlari.add(kart)
    veri['md_kartlari'] = md_kartlari
    veri['atla_kartlari'] = atla_kartlari

    # --- 3.6  Baslangic stoklari  (Stok sayfasi) ---------------------------
    stok = wb['Stok']
    b_otd, b_md, b_ta = {}, {}, {}
    for r in range(3, 17):
        kart = _normalize_kart(stok.cell(row=r, column=1).value)
        if not kart or kart not in KARTLAR:
            continue
        b_otd[kart] = _sayi(stok.cell(row=r, column=2).value)
        b_md[kart] = _sayi(stok.cell(row=r, column=3).value)
        b_ta[kart] = _sayi(stok.cell(row=r, column=4).value)
    veri['baslangic_otd'] = b_otd
    veri['baslangic_md'] = b_md
    veri['baslangic_ta'] = b_ta

    # --- 3.7  Montaj plani  (SUS satir 250-263, sutun D..Q -- R ATILIR) ----
    montaj_plani = {}
    mp0, mp1 = SUS_BLOKLARI['montaj_plani']
    for r in range(mp0, mp1 + 1):
        kart = _normalize_kart(sus.cell(row=r, column=2).value)
        if not kart or kart not in KARTLAR:
            continue
        for g, c in enumerate(range(ILK_GUN_SUTUN, SON_GUN_SUTUN + 1), start=1):
            montaj_plani[(kart, g)] = _sayi(sus.cell(row=r, column=c).value)
    veri['montaj_plani'] = montaj_plani

    # --- 3.8  Donem kadrosu  (SUS satir 285-306) ---------------------------
    kadro = {}
    k0, k1 = SUS_BLOKLARI['kadro']
    for r in range(k0, k1 + 1):
        ad = sus.cell(row=r, column=3).value     # C sutunu = hat/birim adi
        if ad is None:
            continue
        kadro[str(ad).strip()] = int(_sayi(sus.cell(row=r, column=4).value))
    veri['kadro'] = kadro

    return veri


# ===========================================================================
# 4. DOGRULAMA  --  Veri okunduktan sonra hizli tutarlilik kontrolu
# ===========================================================================

def dogrula(veri):
    """Okunan verinin temel tutarliligini kontrol eder, uyarilari listeler."""
    uyarilar = []

    if len(veri['gunler']) != UFUK_GUN_SAYISI:
        uyarilar.append(f"Ufuk {len(veri['gunler'])} gun -- beklenen {UFUK_GUN_SAYISI}.")

    if len(veri['kartlar']) != 14:
        uyarilar.append(f"Kart sayisi {len(veri['kartlar'])} -- beklenen 14.")

    # Alokasyondaki her kart tanimli kartlar arasinda mi?
    for (hat, g), kart in veri['otd_alokasyon'].items():
        if kart not in veri['kartlar']:
            uyarilar.append(f"OTD alokasyonda taninmayan kart: {kart} ({hat}, gun {g}).")

    # Alokasyonda kullanilan her (hat,kart) icin tempo var mi?
    for (hat, g), kart in veri['otd_alokasyon'].items():
        if (hat, kart) not in veri['tempo']:
            uyarilar.append(f"OTD ({hat}, {kart}) icin tempo TANIMSIZ -- Tempolar sayfasini kontrol et.")
    for (hat, g), kart in veri['md_alokasyon'].items():
        # MD hatlari artik kanal (MD2_1, MD2_2); tempo fiziksel hat ile kayitli
        fiz_hat = veri['md_kanal_hat'].get(hat, hat)
        if (fiz_hat, kart) not in veri['tempo']:
            uyarilar.append(f"MD ({hat} -> {fiz_hat}, {kart}) icin tempo TANIMSIZ.")

    # MD'ye giren kartlarin hepsi tempo/alokasyon ile uyumlu mu?
    md_alok_kartlar = {k for k in veri['md_alokasyon'].values()}
    for kart in md_alok_kartlar:
        if kart not in veri['md_kartlari']:
            uyarilar.append(f"{kart} MD'de uretiliyor ama Process_Mapping'te MD=Yok.")

    # Kadrosu 0 olan ama SUS'ta alokasyonu olan hatlar (bilgi amacli)
    for hat in veri['otd_hatlari']:
        if veri['kadro'].get(hat, None) == 0:
            uyarilar.append(f"BILGI: {hat} hattinin kadrosu 0 -- bu donem fiilen uretim yapamaz.")

    return uyarilar


# ===========================================================================
# 5. KOMUT SATIRINDAN CALISTIRMA  --  hizli inceleme / test
# ===========================================================================

if __name__ == '__main__':
    import sys

    dosya = sys.argv[1] if len(sys.argv) > 1 else 'Sasi_Uretim_Plani_v4_1.xlsx'
    print(f"Okunuyor: {dosya}")
    print('=' * 60)

    veri = veri_oku(dosya)

    print(f"Kart sayisi      : {len(veri['kartlar'])}  -> {veri['kartlar']}")
    print(f"Ufuk             : {len(veri['gunler'])} gun "
          f"({veri['gun_tarih'][1]} ... {veri['gun_tarih'][len(veri['gunler'])]})")
    print(f"OTD hatlari      : {veri['otd_hatlari']}")
    print(f"MD hatlari       : {veri['md_hatlari']}")
    print(f"MD'ye giren kart : {sorted(veri['md_kartlari'])}")
    print(f"TA'ya atlayan    : {sorted(veri['atla_kartlari'])}")
    print(f"OTD alokasyon    : {len(veri['otd_alokasyon'])} (hat,gun) hucresi dolu")
    print(f"MD alokasyon     : {len(veri['md_alokasyon'])} (hat,gun) hucresi dolu")
    print(f"Tempo kayitlari  : {len(veri['tempo'])} adet (hat,kart) ciftine tempo")
    print(f"Montaj plani     : {len(veri['montaj_plani'])} (kart,gun) hucresi")

    toplam_montaj = sum(veri['montaj_plani'].values())
    print(f"14 gunluk toplam montaj talebi : {toplam_montaj:.0f}")

    print('-' * 60)
    print("DOGRULAMA SONUCLARI:")
    uyarilar = dogrula(veri)
    if not uyarilar:
        print("  Tutarsizlik bulunamadi.")
    else:
        for u in uyarilar:
            print(f"  - {u}")
